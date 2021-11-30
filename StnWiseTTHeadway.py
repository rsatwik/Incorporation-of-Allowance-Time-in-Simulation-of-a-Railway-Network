#!/usr/bin/env python
# coding: utf-8

# In[126]:


#folder_path = '/home/bhushan/Madhu_Belur/NDLS_MMCT/'       # TraversalDetails from this folder

infraPath = "simulator_input/"                           # All the input files 
simulator_out = "simulator_output/"                      # TD file 
postprocessed_path = "postprocessed_files/"              # Write the file here 

import os
import pathlib
from tqdm import tqdm 

# In[109]:


import pandas as pd
print("Please carefully see which TD to use...")
def min_to_HHMM(minutes):
    day = minutes//1440+1
    HH = (minutes%1440)//60                  # Hours 
    MM = (minutes%1440)%60
    if(len(str(HH))==1):
        HH = '0'+str(HH)
    if(len(str(MM))==1):
        MM = '0'+str(MM)
    return [str(HH)+':'+str(MM),str(day)]

loop_file = open(infraPath+'loop.txt').readlines()[2:]
loop_dict =  {s.split()[0]:s.split()[3] for s in loop_file}
station_file = open(infraPath+'station.txt').readlines()[2:]
station_list = [s.split()[0] for s in station_file]
unsch = open(infraPath+'unscheduled.txt').readlines()[2:]
direction_dict = {u.split()[0]:u.split()[1] for u in unsch}
priority_dict = {u.split()[0]:u.split()[6] for u in unsch}
train_nos_up = []
train_nos_down = []

for t in direction_dict:
    if(direction_dict[t]=='up'):
        train_nos_up.append(t)
    else:
        train_nos_down.append(t)

TD_file = open(simulator_out+'TraversalDetails.txt').readlines()
TDDc = {}
for i,t in enumerate(TD_file):
    if('Printing timetables for train' in t):
        if(i!=0):
            TDDc[curr_train] = dic
            
        dic = {}
        curr_train = t.split()[-1]
    else:
        if(t.split()[2] in loop_dict):
            dic[loop_dict[t.split()[2]]]=[t.split()[11],t.split()[14]]
TDDc[curr_train] = dic

df_list = []
df_list_min = []
for s in station_list:
    d = {}
    d_list = []
    for train in TDDc:
        
        if(s in TDDc[train]):
            d = {'Train':str(train),'Station':s,'Arrival':TDDc[train][s][0],'Departure':TDDc[train][s][1],'Day':int(float(TDDc[train][s][0]))//1440+1}
            d_list.append(d)
    df = pd.DataFrame.from_dict(d_list)
    
    df['Arrival'] = df['Arrival'].astype(float)
    df = df.sort_values('Arrival')
    df_min = df.copy()                               # Arrival and Departure Times in minutes 
    df['Train'] = df['Train'].astype(str)
    df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(x))[0])
    df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
    df_list.append(df)
    df_list_min.append(df_min)


# In[111]:


# Basically write to an Excel sheet 
from openpyxl import Workbook

book = Workbook()
sheet = book.active

row_index = 2
column_index = 2
for d in tqdm(df_list):
    train_nos = list(d['Train'])
    station = list(d['Station'])[0]                      # Station
    sheet.cell(row=row_index-1, column=column_index-1).value = station          # Column 2 
    sheet.cell(row=row_index, column=column_index).value = 'Train'          # Column 2 
    sheet.cell(row=row_index, column=column_index+1).value = 'Priority'          # Column 2
    sheet.cell(row=row_index, column=column_index+2).value = 'Arrival'          # Column 2
    sheet.cell(row=row_index, column=column_index+3).value = 'Departure'          # Column 2
    sheet.cell(row=row_index, column=column_index+4).value = 'Day'          # Column 2
    row_index+=1
    
    
    
    for t in train_nos:
        if(t not in direction_dict):
            continue
        if(direction_dict[t]=='up'):
            sheet.cell(row=row_index, column=column_index).value = t
            sheet.cell(row=row_index, column=column_index+1).value = str(priority_dict[t])          # Column 2
            sheet.cell(row=row_index, column=column_index+2).value = str(d.loc[d['Train']==t, 'Arrival'].values[0])          # Arrival
            sheet.cell(row=row_index, column=column_index+3).value = str(d.loc[d['Train']==t, 'Departure'].values[0])   
            sheet.cell(row=row_index, column=column_index+4).value = str(d.loc[d['Train']==t, 'Day'].values[0])
            row_index+=1
            
        else:
            continue
        
    
    row_index = 2
    column_index+=7

print("Creating a new directory...")
pathlib.Path(postprocessed_path+'Stationwise_Headway').mkdir( exist_ok=True)


book.save(postprocessed_path+'Stationwise_Headway/'+'Stnwisearrdepup.xlsx')


# In[112]:


# Basically write to an Excel sheet 
from openpyxl import Workbook

book = Workbook()
sheet = book.active

row_index = 2
column_index = 2
for d in tqdm(df_list):
    train_nos = list(d['Train'])
    station = list(d['Station'])[0]                      # Station
    sheet.cell(row=row_index-1, column=column_index-1).value = station          # Column 2 
    sheet.cell(row=row_index, column=column_index).value = 'Train'          # Column 2 
    sheet.cell(row=row_index, column=column_index+1).value = 'Priority'          # Column 2
    sheet.cell(row=row_index, column=column_index+2).value = 'Arrival'          # Column 2
    sheet.cell(row=row_index, column=column_index+3).value = 'Departure'          # Column 2
    sheet.cell(row=row_index, column=column_index+4).value = 'Day'          # Column 2
    row_index+=1
    
    
    
    for t in train_nos:
        if(t not in direction_dict):
            continue
        if(direction_dict[t]=='down'):
            sheet.cell(row=row_index, column=column_index).value = t
            sheet.cell(row=row_index, column=column_index+1).value = str(priority_dict[t])          # Column 2
            sheet.cell(row=row_index, column=column_index+2).value = str(d.loc[d['Train']==t, 'Arrival'].values[0])          # Arrival
            sheet.cell(row=row_index, column=column_index+3).value = str(d.loc[d['Train']==t, 'Departure'].values[0])   
            sheet.cell(row=row_index, column=column_index+4).value = str(d.loc[d['Train']==t, 'Day'].values[0])
            row_index+=1
            
        else:
            continue
        
    
    row_index = 2
    column_index+=7




book.save(postprocessed_path+'Stationwise_Headway/'+'Stnwisearrdepdown.xlsx')


# In[113]:


df_list_up = []
df_list_down = []                # Minutes list 
for d in df_list_min:
    df_1 = d[d['Train'].isin(train_nos_up)]
    df_2 = d[d['Train'].isin(train_nos_down)]
    df_list_up.append(df_1)
    df_list_down.append(df_2)
    
# Do this for up and down direction trains separately 
# 7 and 10 minute headway separately 
# dic_lis = {}
# dic = {}
# df_headway = []
df = pd.DataFrame(columns = list(d.columns))
for d in tqdm(df_list_up):
    d = d.reset_index()
    del d['index']
    d['Arr_dif'] = d['Arrival'].diff()
    index = list(d[d['Arr_dif']<7].index)
#     print(index)
    
    if(not index):
        continue
    else:
        for k,i in enumerate(index):
            df = df.append(d.iloc[i-1],ignore_index=True)
            df = df.append(d.iloc[i],ignore_index=True)
#             df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
#             df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
            df.reset_index()
            
#             pd.concat([df,d.iloc[i-1]],axis=1)
#             pd.concat([df,d.iloc[i]],axis=1)

for i in list(df.index):
    if(i%2==0):
        df.iloc[i,5] = '-'
    
df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])        
df.to_excel(postprocessed_path+'Stationwise_Headway/'+'UpDirArrHeadway.xlsx',index=False)
df = pd.DataFrame(columns = list(d.columns))
for d in tqdm(df_list_down):
    d = d.reset_index()
    del d['index']
    d['Arr_dif'] = d['Arrival'].diff()
    index = list(d[d['Arr_dif']<7].index)
#     print(index)
    
    if(not index):
        continue
    else:
        for k,i in enumerate(index):
            df = df.append(d.iloc[i-1],ignore_index=True)
            df = df.append(d.iloc[i],ignore_index=True)
#             df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
#             df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
            df.reset_index()
            
#             pd.concat([df,d.iloc[i-1]],axis=1)
#             pd.concat([df,d.iloc[i]],axis=1)

for i in list(df.index):
    if(i%2==0):
        df.iloc[i,5] = '-'
df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0]) 
df.to_excel(postprocessed_path+'Stationwise_Headway/'+'DownDirArrHeadway.xlsx',index=False)



# First you have to sort according to departure 
df = pd.DataFrame(columns = list(d.columns))
for d in tqdm(df_list_up):
    d['Departure'] = d['Departure'].astype(float)
    d = d.sort_values('Departure')
    d = d.reset_index()
    del d['index']
    d['Dep_diff'] = d['Departure'].diff()
    index = list(d[d['Dep_diff']<7].index)
#     print(index)
    
    if(not index):
        continue
    else:
        for k,i in enumerate(index):
            df = df.append(d.iloc[i-1],ignore_index=True)
            df = df.append(d.iloc[i],ignore_index=True)
#             df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
#             df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
            df.reset_index()
            
#             pd.concat([df,d.iloc[i-1]],axis=1)
#             pd.concat([df,d.iloc[i]],axis=1)
del df['Arr_dif']         # Delete this 
for i in list(df.index):
    if(i%2==0):
        df.iloc[i,5] = '-'
    
df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])        
df.to_excel(postprocessed_path+'Stationwise_Headway/'+'UpDirDepHeadway.xlsx',index=False)

df = pd.DataFrame(columns = list(d.columns))
for d in tqdm(df_list_down):
    d['Departure'] = d['Departure'].astype(float)
    d = d.sort_values('Departure')
    d = d.reset_index()
    del d['index']
    d['Dep_diff'] = d['Departure'].diff()
    index = list(d[d['Dep_diff']<7].index)
#     print(index)
    
    if(not index):
        continue
    else:
        for k,i in enumerate(index):
            df = df.append(d.iloc[i-1],ignore_index=True)
            df = df.append(d.iloc[i],ignore_index=True)
#             df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
#             df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
            df.reset_index()
            
#             pd.concat([df,d.iloc[i-1]],axis=1)
#             pd.concat([df,d.iloc[i]],axis=1)

for i in list(df.index):
    if(i%2==0):
        df.iloc[i,5] = '-'
df['Arrival'] = df['Arrival'].apply(lambda x: min_to_HHMM(int(float(x)))[0])
df['Departure'] = df['Departure'].apply(lambda x: min_to_HHMM(int(float(x)))[0]) 
df.to_excel(postprocessed_path+'Stationwise_Headway/'+'DownDirDepHeadway.xlsx',index=False)


# In[114]:


# Combine all the generated sheets now:
import pandas as pd
from collections import defaultdict
import os
from tqdm import tqdm
from pandas import ExcelWriter
import glob
import sys
import gspread
from oauth2client.service_account import ServiceAccountCredentials

writer = ExcelWriter(postprocessed_path+"Headway_Stationwise_TT.xlsx")
# file_list = [postprocessPath+'{}-{}.xlsx'.format(s,reference_stations_down[reference_stations_down.index(s)+1]) for s in reference_stations_down if s!=reference_stations_down[-1]]
for filename in tqdm(glob.glob(postprocessed_path+'Stationwise_Headway/'+"*.xlsx")):                     # But combining all the sheets maybe wrong 
#     if(filename not in file_list):
#         continue

    excel_file = pd.ExcelFile(filename)
    (_, f_name) = os.path.split(filename)
    (f_short_name, _) = os.path.splitext(f_name)
    for sheet_name in excel_file.sheet_names:
        df_excel = pd.read_excel(filename)
        df_excel.to_excel(writer, f_short_name, index=False)

writer.save()
